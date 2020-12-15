# -*- coding: UTF-8 -*-

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

query = "Machine learning skills"
query = query.replace(' ', '+') 

wb = Workbook()             # open new workbook, use load_workbook if existing
ws = wb.create_sheet(title="Google search")
ws.append(["Title","Link"])

for page in range(3):
    URL = 'https://google.com/search?q=' + query + "&start=" + str(page)
    # desktop user-agent
    USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.14; rv:65.0) Gecko/20100101 Firefox/65.0"
    headers = {"user-agent" : USER_AGENT}
    resp = requests.get(URL, headers=headers)
    if resp.status_code == 200:
        soup = BeautifulSoup(resp.content, "html.parser")
        results = []
        for g in soup.find_all('div', class_='rc'):  
            title = g.find('h3').text
            link = g.a.get('href')
            item = {
                "title": title,
                "link": link
                }
            results.append(item)
            row = [title, link]  
            ws.append(row)            
        wb.save('google_search.xlsx')
        print(results)  





